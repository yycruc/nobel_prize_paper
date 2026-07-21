from __future__ import annotations

import csv
import json
import statistics
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
ANALYSIS_DIR = ROOT / "02_full_collection" / "06_analysis"
REPORT_DIR = ROOT / "02_full_collection" / "08_report"
OUT_XLSX = REPORT_DIR / "nobel_report_figures.xlsx"
OUT_JSON = REPORT_DIR / "nobel_report_metrics.json"


COUNTRIES = ["United States", "China", "United Kingdom", "Germany", "Japan", "France"]
COUNTRY_CN = {
    "United States": "美国",
    "China": "中国",
    "United Kingdom": "英国",
    "Germany": "德国",
    "Japan": "日本",
    "France": "法国",
}


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        return list(csv.DictReader(fh))


def decade_sort_key(label: str) -> int:
    return int(label[:4])


def add_table(ws, name: str, rows: list[dict[str, Any]], headers: list[str]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    if rows:
        ref = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"
        table = Table(displayName=name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in ws.columns:
        width = 14
        if col[0].value in {"decade", "journal", "metric", "note"}:
            width = 28
        ws.column_dimensions[col[0].column_letter].width = width


def pdf_award_rows() -> list[dict[str, Any]]:
    columns = ["美国", "英国", "德国", "法国", "日本", "瑞士", "瑞典", "荷兰", "俄国", "加拿大", "其他", "合计"]
    raw = [
        ("1901-1910", [1, 5, 12, 6, 0, 1, 1, 4, 2, 0, 4, 36]),
        ("1911-1920", [1, 3, 7, 4, 0, 2, 2, 1, 0, 0, 4, 24]),
        ("1921-1930", [2, 7, 9, 3, 0, 0, 2, 2, 0, 2, 6, 33]),
        ("1931-1940", [9, 7, 8, 2, 0, 2, 0, 1, 0, 0, 6, 35]),
        ("1941-1950", [15, 6, 3, 0, 1, 3, 1, 0, 0, 0, 7, 36]),
        ("1951-1960", [28, 9, 3, 0, 0, 0, 1, 1, 4, 0, 6, 52]),
        ("1961-1970", [27, 11, 5, 5, 1, 0, 3, 0, 3, 0, 4, 59]),
        ("1971-1980", [38, 12, 3, 1, 1, 2, 0, 0, 1, 1, 8, 67]),
        ("1981-1990", [36, 4, 9, 1, 2, 2, 4, 1, 0, 3, 2, 64]),
        ("1991-2000", [38, 4, 5, 3, 1, 2, 1, 3, 1, 2, 3, 63]),
        ("2001-2010", [37, 12, 5, 4, 8, 1, 0, 1, 2, 0, 6, 76]),
        ("2011-2020", [33, 14, 4, 5, 8, 3, 1, 1, 0, 2, 8, 79]),
    ]
    return [{"decade": decade, **dict(zip(columns, values))} for decade, values in raw]


def lag_rows() -> list[dict[str, Any]]:
    rows = read_csv(ANALYSIS_DIR / "q1_lag_by_award_decade.csv")
    return [
        {
            "award_decade": row["award_decade"],
            "mean_lag_years": float(row["mean_lag_years"]),
            "median_lag_years": float(row["median_lag_years"]),
            "key_paper_rows": int(row["key_paper_rows"]),
        }
        for row in sorted(rows, key=lambda r: decade_sort_key(r["award_decade"]))
    ]


def lag_by_category_rows() -> list[dict[str, Any]]:
    rows = read_csv(ANALYSIS_DIR / "analysis_ready_key_papers_full.csv")
    groups: dict[tuple[str, str], list[int]] = defaultdict(list)
    overall: dict[str, list[int]] = defaultdict(list)
    for row in rows:
        year = int(row["award_year"])
        dec = f"{year // 10 * 10}s"
        lag = int(row["award_lag_years"])
        cat = row["category"]
        groups[(dec, cat)].append(lag)
        overall[dec].append(lag)
    out = []
    for dec in sorted(overall, key=decade_sort_key):
        out.append(
            {
                "award_decade": dec,
                "总体": round(statistics.median(overall[dec]), 2),
                "物理学": round(statistics.median(groups.get((dec, "Physics"), [])), 2) if groups.get((dec, "Physics")) else "",
                "化学": round(statistics.median(groups.get((dec, "Chemistry"), [])), 2) if groups.get((dec, "Chemistry")) else "",
                "生理学或医学": round(statistics.median(groups.get((dec, "Physiology or Medicine"), [])), 2) if groups.get((dec, "Physiology or Medicine")) else "",
            }
        )
    return out


def journal_coverage_rows() -> list[dict[str, Any]]:
    rows = read_csv(ANALYSIS_DIR / "additional_journal_coverage_rank.csv")
    out = []
    for row in rows:
        rank = int(row["rank"])
        if rank <= 77:
            out.append(
                {
                    "rank": rank,
                    "journal": row["journal"],
                    "key_paper_rows": int(row["key_paper_rows"]),
                    "cumulative_row_share": float(row["cumulative_row_share"]),
                    "coverage_90_line": 0.9,
                }
            )
    return out


def share_rows_top77() -> list[dict[str, Any]]:
    rows = read_csv(ANALYSIS_DIR / "additional_country_year_world_shares_1850_aggregate.csv")
    out = []
    for row in rows:
        country = row["country_name"]
        year = int(row["year"])
        if country in COUNTRIES and year >= 1980:
            out.append(
                {
                    "year": year,
                    "country": COUNTRY_CN[country],
                    "country_en": country,
                    "country_count": int(row["selected_journal_country_count"]),
                    "world_count": int(row["selected_journal_world_count"]),
                    "share": float(row["country_world_share"]),
                }
            )
    return out


def aggregate_top20() -> list[dict[str, Any]]:
    rows = read_csv(ANALYSIS_DIR / "additional_country_journal_year_world_shares_1850.csv")
    totals: dict[tuple[int, str], list[int]] = defaultdict(lambda: [0, 0])
    for row in rows:
        rank = int(row["journal_rank"])
        country = row["country_name"]
        if rank <= 20 and country in COUNTRIES:
            key = (int(row["year"]), country)
            totals[key][0] += int(row["country_publication_count"])
            totals[key][1] += int(row["world_publication_count"])
    out = []
    for (year, country), (country_count, world_count) in sorted(totals.items()):
        out.append(
            {
                "year": year,
                "country": COUNTRY_CN[country],
                "country_en": country,
                "country_count": country_count,
                "world_count": world_count,
                "share": country_count / world_count if world_count else 0,
            }
        )
    return out


def first_threshold_year(rows: list[dict[str, Any]], country_cn: str, threshold: float) -> int | None:
    points = sorted([r for r in rows if r["country"] == country_cn], key=lambda r: r["year"])
    trailing: list[float] = []
    for row in points:
        trailing.append(row["share"])
        if len(trailing) > 5:
            trailing.pop(0)
        if len(trailing) == 5 and statistics.mean(trailing) >= threshold:
            return int(row["year"])
    return None


def threshold_rows(top77: list[dict[str, Any]], top20: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out = []
    for country in ["美国", "中国", "日本", "英国", "德国", "法国"]:
        for threshold in [0.01, 0.05, 0.10]:
            out.append(
                {
                    "country": country,
                    "threshold": f"{int(threshold * 100)}%",
                    "top77_year": first_threshold_year(top77, country, threshold) or "",
                    "top20_year": first_threshold_year(top20, country, threshold) or "",
                }
            )
    return out


def wide_country_share(rows: list[dict[str, Any]], start_year: int = 1980) -> list[dict[str, Any]]:
    by_year: dict[int, dict[str, float]] = defaultdict(dict)
    for row in rows:
        year = int(row["year"])
        if year >= start_year:
            by_year[year][row["country"]] = row["share"]
    return [{"year": year, **{c: by_year[year].get(c, 0) for c in ["美国", "中国", "日本", "英国", "德国", "法国"]}} for year in sorted(by_year)]


def china_compare_rows(top77: list[dict[str, Any]], top20: list[dict[str, Any]]) -> list[dict[str, Any]]:
    top77_by_year = {int(row["year"]): row["share"] for row in top77 if row["country"] == "中国"}
    top20_by_year = {int(row["year"]): row["share"] for row in top20 if row["country"] == "中国"}
    out = []
    for year in range(2000, 2026):
        out.append({"year": year, "Top77": top77_by_year.get(year, 0), "Top20": top20_by_year.get(year, 0)})
    return out


def metric_summary() -> dict[str, Any]:
    analysis = read_csv(ANALYSIS_DIR / "analysis_ready_key_papers_full.csv")
    basic = {row["metric"]: row["value"] for row in read_csv(ANALYSIS_DIR / "additional_basic_stats.csv")}
    top10 = read_csv(ANALYSIS_DIR / "q2_top10_key_paper_journals.csv")
    top77_share = read_csv(ANALYSIS_DIR / "additional_journal_coverage_rank.csv")[76]
    lags = [int(row["award_lag_years"]) for row in analysis if row.get("award_lag_years")]
    return {
        "baseline_records": int(float(basic["all_natural_science_nobel_records"])),
        "main_records": int(float(basic["records_with_analysis_ready_key_papers"])),
        "main_rows": len(analysis),
        "unique_papers": int(float(basic["unique_key_papers"])),
        "unique_journals": int(float(basic["unique_journals"])),
        "top77_cumulative_share": float(top77_share["cumulative_row_share"]),
        "top10": [
            {
                "rank": int(row["rank"]),
                "journal": row["journal"],
                "key_paper_rows": int(row["key_paper_rows"]),
                "covered_records": int(row["covered_nobel_records"]),
            }
            for row in top10
        ],
        "lag_mean": round(statistics.mean(lags), 2),
        "lag_median": round(statistics.median(lags), 2),
        "lag_max": max(lags),
    }


def build() -> None:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    awards = pdf_award_rows()
    lag = lag_rows()
    lag_cat = lag_by_category_rows()
    coverage = journal_coverage_rows()
    top77 = share_rows_top77()
    top20 = aggregate_top20()
    thresholds = threshold_rows(top77, top20)
    top77_wide = wide_country_share(top77)
    top20_wide = wide_country_share(top20)
    china_compare = china_compare_rows(top77, top20)
    metrics = metric_summary()
    metrics["thresholds"] = thresholds

    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    ws["A1"] = "Nobel report figures"
    ws["A1"].font = Font(name="Microsoft YaHei", bold=True, size=16)
    ws["A3"] = "Charts are added and exported by add_report_excel_charts.ps1."
    ws.column_dimensions["A"].width = 90
    ws.sheet_view.showGridLines = False

    add_table(wb.create_sheet("PDF_AwardsByDecade"), "tblAwardsByDecade", awards, ["decade", "美国", "英国", "德国", "法国", "日本", "瑞士", "瑞典", "荷兰", "俄国", "加拿大", "其他", "合计"])
    add_table(wb.create_sheet("LagTrend"), "tblLagTrend", lag, ["award_decade", "mean_lag_years", "median_lag_years", "key_paper_rows"])
    add_table(wb.create_sheet("LagByCategory"), "tblLagByCategory", lag_cat, ["award_decade", "总体", "物理学", "化学", "生理学或医学"])
    add_table(wb.create_sheet("JournalCoverage"), "tblJournalCoverageReport", coverage, ["rank", "journal", "key_paper_rows", "cumulative_row_share", "coverage_90_line"])
    add_table(wb.create_sheet("Top77CountryShare"), "tblTop77CountryShare", top77_wide, ["year", "美国", "中国", "日本", "英国", "德国", "法国"])
    add_table(wb.create_sheet("Top20CountryShare"), "tblTop20CountryShare", top20_wide, ["year", "美国", "中国", "日本", "英国", "德国", "法国"])
    add_table(wb.create_sheet("ChinaTop77Top20"), "tblChinaTop77Top20", china_compare, ["year", "Top77", "Top20"])
    add_table(wb.create_sheet("ThresholdYears"), "tblThresholdYears", thresholds, ["country", "threshold", "top77_year", "top20_year"])

    wb.save(OUT_XLSX)
    OUT_JSON.write_text(json.dumps(metrics, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"xlsx": str(OUT_XLSX), "metrics": str(OUT_JSON)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    build()
