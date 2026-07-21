from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
ANALYSIS_CSV = ROOT / "02_full_collection" / "06_analysis" / "analysis_ready_key_papers_full.csv"
Q2_CSV = ROOT / "02_full_collection" / "06_analysis" / "q2_top10_key_paper_journals.csv"
Q3_PANEL_CSV = ROOT / "02_full_collection" / "06_analysis" / "q3_country_journal_year_counts_top10.csv"
Q3_AGG_CSV = ROOT / "02_full_collection" / "06_analysis" / "q3_country_year_counts_top10_aggregate.csv"
ADD_BASIC_CSV = ROOT / "02_full_collection" / "06_analysis" / "additional_basic_stats.csv"
ADD_DIST_CSV = ROOT / "02_full_collection" / "06_analysis" / "additional_record_paper_count_distribution.csv"
ADD_COVERAGE_CSV = ROOT / "02_full_collection" / "06_analysis" / "additional_journal_coverage_rank.csv"
ADD_PERIOD_CSV = ROOT / "02_full_collection" / "06_analysis" / "additional_journal_coverage_by_award_period.csv"
ADD_SHARE_PANEL_CSV = ROOT / "02_full_collection" / "06_analysis" / "additional_country_journal_year_world_shares_1850.csv"
ADD_SHARE_AGG_CSV = ROOT / "02_full_collection" / "06_analysis" / "additional_country_year_world_shares_1850_aggregate.csv"
OUT_DIR = ROOT / "02_full_collection" / "07_excel"
BASE_XLSX = OUT_DIR / "nobel_key_papers_pivot_base.xlsx"


def decade(year: int) -> str:
    return f"{year // 10 * 10}s"


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        return list(csv.DictReader(fh))


def add_table(ws, name: str, headers: list[str], rows: list[dict[str, object]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    ref = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"
    table = Table(displayName=name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for col in ws.columns:
        header = col[0].value
        width = 14
        if header in {"title", "journal", "full_name", "openalex_source_id", "openalex_work_id", "doi", "metric"}:
            width = 28
        if header in {"reference_text", "notes", "evidence_sources", "country_count_filter", "world_count_filter", "count_filter"}:
            width = 36
        if header in {"row_share", "cumulative_row_share", "share_of_all_nobel_records", "country_world_share"}:
            width = 16
        ws.column_dimensions[col[0].column_letter].width = width
    ws.sheet_view.showGridLines = False


def build() -> Path:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    readme = wb.active
    readme.title = "Readme"
    readme["A1"] = "Nobel Key Papers: Data, Pivot Tables, and Charts"
    readme["A1"].font = Font(bold=True, size=16)
    readme["A3"] = "Workbook contents"
    readme["A3"].font = Font(bold=True)
    readme["A4"] = "Data_KeyPapers: main analysis table, 823 rows covering 504 Nobel records."
    readme["A5"] = "Data_Top10Journals: top 10 journals by Nobel key-paper rows."
    readme["A6"] = "Data_CountryJournalYear: 10 journals x 6 countries x 1880-2025 yearly article-count panel."
    readme["A7"] = "Data_CountryYearAgg: six-country annual totals across the top 10 journals."
    readme["A8"] = "Q1/Q2/Q3 sheets are added by Excel COM as native PivotTables and charts."
    readme["A9"] = "Data_BasicStats / Data_RecordPaperDist / Data_JournalCoverage / Data_JournalPeriod: additional coverage and distribution tables."
    readme["A10"] = "Data_CountryWorldShare / Data_CountryWorldAgg: selected 90%-coverage journal country shares, 1850-2025, using OpenAlex world totals as denominator."
    readme["A11"] = "Q4/Q5/Q6 sheets are added by Excel COM as additional native PivotTables and charts."
    readme["A13"] = "Scope notes"
    readme["A13"].font = Font(bold=True)
    readme["A14"] = "Main analysis includes rows with title, year, journal/source, and auditable metadata matching. Lower-confidence or key-relevance-review candidates are excluded from the main statistics."
    readme["A15"] = "Q3 country article counts use OpenAlex top 10 journal sources, authorships.countries, and type:article."
    readme["A16"] = "Q6 country shares use the 77 journals required to cover 90.0% of Nobel key-paper rows; country shares are country article count divided by all-world article count in the same journal set and year."
    readme.column_dimensions["A"].width = 120
    for row in range(1, 17):
        readme[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
    readme.sheet_view.showGridLines = False

    key_rows = []
    for row in read_csv(ANALYSIS_CSV):
        award_year = int(row["award_year"])
        publication_year = int(row["publication_year"])
        key_rows.append(
            {
                "analysis_id": row["analysis_id"],
                "registry_id": row["registry_id"],
                "validation_id": row["validation_id"],
                "laureate_id": row["laureate_id"],
                "full_name": row["full_name"],
                "award_year": award_year,
                "award_decade": decade(award_year),
                "category": row["category"],
                "title": row["title"],
                "publication_year": publication_year,
                "publication_decade": decade(publication_year),
                "award_lag_years": int(row["award_lag_years"]),
                "journal": row["journal"],
                "openalex_source_id": row["openalex_source_id"],
                "issn_l": row["issn_l"],
                "doi": row["doi"],
                "openalex_work_id": row["openalex_work_id"],
                "evidence_sources": row["evidence_sources"],
                "analysis_role": row["analysis_role"],
            }
        )
    add_table(
        wb.create_sheet("Data_KeyPapers"),
        "tblKeyPapers",
        ["analysis_id", "registry_id", "validation_id", "laureate_id", "full_name", "award_year", "award_decade", "category", "title", "publication_year", "publication_decade", "award_lag_years", "journal", "openalex_source_id", "issn_l", "doi", "openalex_work_id", "evidence_sources", "analysis_role"],
        key_rows,
    )

    top_rows = []
    for row in read_csv(Q2_CSV):
        top_rows.append(
            {
                "rank": int(row["rank"]),
                "journal": row["journal"],
                "key_paper_rows": int(row["key_paper_rows"]),
                "covered_nobel_records": int(row["covered_nobel_records"]),
                "openalex_source_id": row["openalex_source_id"],
                "issn_l": row["issn_l"],
            }
        )
    add_table(
        wb.create_sheet("Data_Top10Journals"),
        "tblTopJournals",
        ["rank", "journal", "key_paper_rows", "covered_nobel_records", "openalex_source_id", "issn_l"],
        top_rows,
    )

    panel_rows = []
    for row in read_csv(Q3_PANEL_CSV):
        year = int(row["year"])
        panel_rows.append(
            {
                "journal_rank": int(row["journal_rank"]),
                "journal": row["journal"],
                "openalex_source_id": row["openalex_source_id"],
                "issn_l": row["issn_l"],
                "country_code": row["country_code"],
                "country_name": row["country_name"],
                "year": year,
                "year_decade": decade(year),
                "publication_count": int(row["publication_count"]),
            }
        )
    add_table(
        wb.create_sheet("Data_CountryJournalYear"),
        "tblCountryJournalYear",
        ["journal_rank", "journal", "openalex_source_id", "issn_l", "country_code", "country_name", "year", "year_decade", "publication_count"],
        panel_rows,
    )

    agg_rows = []
    for row in read_csv(Q3_AGG_CSV):
        year = int(row["year"])
        agg_rows.append(
            {
                "country_code": row["country_code"],
                "country_name": row["country_name"],
                "year": year,
                "year_decade": decade(year),
                "top10_journal_publication_count": int(row["top10_journal_publication_count"]),
            }
        )
    add_table(
        wb.create_sheet("Data_CountryYearAgg"),
        "tblCountryYearAgg",
        ["country_code", "country_name", "year", "year_decade", "top10_journal_publication_count"],
        agg_rows,
    )

    basic_rows = []
    for row in read_csv(ADD_BASIC_CSV):
        value = row["value"]
        try:
            value = float(value) if "." in str(value) else int(value)
        except ValueError:
            pass
        basic_rows.append({"metric": row["metric"], "value": value, "notes": row["notes"]})
    add_table(wb.create_sheet("Data_BasicStats"), "tblBasicStats", ["metric", "value", "notes"], basic_rows)

    dist_rows = []
    for row in read_csv(ADD_DIST_CSV):
        dist_rows.append(
            {
                "paper_count_per_nobel_record": int(row["paper_count_per_nobel_record"]),
                "nobel_records": int(row["nobel_records"]),
                "share_of_all_nobel_records": float(row["share_of_all_nobel_records"]),
            }
        )
    add_table(wb.create_sheet("Data_RecordPaperDist"), "tblRecordPaperDist", ["paper_count_per_nobel_record", "nobel_records", "share_of_all_nobel_records"], dist_rows)

    coverage_rows = []
    for row in read_csv(ADD_COVERAGE_CSV):
        coverage_rows.append(
            {
                "rank": int(row["rank"]),
                "journal": row["journal"],
                "key_paper_rows": int(row["key_paper_rows"]),
                "row_share": float(row["row_share"]),
                "cumulative_key_paper_rows": int(row["cumulative_key_paper_rows"]),
                "cumulative_row_share": float(row["cumulative_row_share"]),
                "covered_nobel_records": int(row["covered_nobel_records"]),
                "cumulative_covered_nobel_records": int(row["cumulative_covered_nobel_records"]),
                "cumulative_record_share_among_covered_records": float(row["cumulative_record_share_among_covered_records"]),
                "openalex_source_id": row["openalex_source_id"],
                "issn_l": row["issn_l"],
                "selected_for_90pct_row_panel": int(row["selected_for_90pct_row_panel"]),
            }
        )
    add_table(
        wb.create_sheet("Data_JournalCoverage"),
        "tblJournalCoverage",
        ["rank", "journal", "key_paper_rows", "row_share", "cumulative_key_paper_rows", "cumulative_row_share", "covered_nobel_records", "cumulative_covered_nobel_records", "cumulative_record_share_among_covered_records", "openalex_source_id", "issn_l", "selected_for_90pct_row_panel"],
        coverage_rows,
    )

    period_rows = []
    for row in read_csv(ADD_PERIOD_CSV):
        period_rows.append(
            {
                "award_period": row["award_period"],
                "rank_within_period": int(row["rank_within_period"]),
                "journal": row["journal"],
                "key_paper_rows": int(row["key_paper_rows"]),
                "period_key_paper_rows": int(row["period_key_paper_rows"]),
                "period_row_share": float(row["period_row_share"]),
                "period_cumulative_row_share_top15": float(row["period_cumulative_row_share_top15"]),
            }
        )
    add_table(
        wb.create_sheet("Data_JournalPeriod"),
        "tblJournalPeriod",
        ["award_period", "rank_within_period", "journal", "key_paper_rows", "period_key_paper_rows", "period_row_share", "period_cumulative_row_share_top15"],
        period_rows,
    )

    share_rows = []
    for row in read_csv(ADD_SHARE_PANEL_CSV):
        share_rows.append(
            {
                "journal_rank": int(row["journal_rank"]),
                "journal": row["journal"],
                "openalex_source_id": row["openalex_source_id"],
                "issn_l": row["issn_l"],
                "country_code": row["country_code"],
                "country_name": row["country_name"],
                "year": int(row["year"]),
                "year_decade": row["year_decade"],
                "country_publication_count": int(row["country_publication_count"]),
                "world_publication_count": int(row["world_publication_count"]),
                "country_world_share": float(row["country_world_share"]),
            }
        )
    add_table(
        wb.create_sheet("Data_CountryWorldShare"),
        "tblCountryWorldShare",
        ["journal_rank", "journal", "openalex_source_id", "issn_l", "country_code", "country_name", "year", "year_decade", "country_publication_count", "world_publication_count", "country_world_share"],
        share_rows,
    )

    share_agg_rows = []
    for row in read_csv(ADD_SHARE_AGG_CSV):
        share_agg_rows.append(
            {
                "country_code": row["country_code"],
                "country_name": row["country_name"],
                "year": int(row["year"]),
                "year_decade": row["year_decade"],
                "selected_journal_country_count": int(row["selected_journal_country_count"]),
                "selected_journal_world_count": int(row["selected_journal_world_count"]),
                "country_world_share": float(row["country_world_share"]),
            }
        )
    add_table(
        wb.create_sheet("Data_CountryWorldAgg"),
        "tblCountryWorldAgg",
        ["country_code", "country_name", "year", "year_decade", "selected_journal_country_count", "selected_journal_world_count", "country_world_share"],
        share_agg_rows,
    )

    wb.save(BASE_XLSX)
    return BASE_XLSX


if __name__ == "__main__":
    print(build())
